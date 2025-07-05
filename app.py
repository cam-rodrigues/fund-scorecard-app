
import streamlit as st
import io
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rapidfuzz import process, fuzz
import base64
import string
from datetime import datetime
import gc

# --- Theme Control ---
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False

# --- Sidebar Navigation ---
st.set_page_config(page_title="FidSync", layout="wide")
st.sidebar.title("FidSync")
st.sidebar.markdown("---")
st.sidebar.subheader("Navigate")
page = st.sidebar.radio("", ["About FidSync", "How to Use"], key="nav")
st.sidebar.markdown("---")
st.sidebar.subheader("Tools")
tool = st.sidebar.radio("", ["Fund Scorecard"], key="tool")
st.sidebar.markdown("---")
st.sidebar.checkbox("Dark Mode (beta)", key="dark_mode")

# === Shared Fill Styles ===
GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

# === Normalization ===
def normalize_name(name):
    name = name.lower().translate(str.maketrans('', '', string.punctuation))
    return " ".join(name.split())

# === PDF Extraction ===
def extract_fund_status(pdf_bytes, start_page, end_page):
    fund_status = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num in range(start_page-1, end_page):  # Adjust to real page numbers
            if page_num >= len(pdf.pages):
                continue
            lines = (pdf.pages[page_num].extract_text() or "").split('\n')
            for i, line in enumerate(lines):
                if "manager tenure" in line.lower() and i > 0:
                    fund_line = lines[i - 1].strip()
                    if "Fund Meets Watchlist Criteria." in fund_line:
                        name = fund_line.split("Fund Meets Watchlist Criteria.")[0].strip()
                        fund_status[normalize_name(name)] = "Pass"
                    elif "Fund has been placed on watchlist" in fund_line:
                        name = fund_line.split("Fund has been placed on watchlist")[0].strip()
                        fund_status[normalize_name(name)] = "Fail"
    return fund_status

# === Excel Update Logic ===
def update_excel_with_status(pdf_bytes, excel_bytes, sheet_name, status_col, start_row, fund_names, start_page, end_page, dry_run=False):
    fund_status_map = extract_fund_status(pdf_bytes, start_page, end_page)
    pdf_names = list(fund_status_map.keys())
    wb = load_workbook(io.BytesIO(excel_bytes))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]

    updated_count = 0
    match_log = []

    for i, fund_name in enumerate(fund_names):
        normalized = normalize_name(fund_name)
        match = process.extractOne(normalized, pdf_names, scorer=fuzz.token_sort_ratio)
        row = start_row + i
        if match and match[1] >= 70:
            matched_name, score = match
            status = fund_status_map.get(matched_name)
            if not dry_run:
                cell = ws[f"{status_col}{row}"]
                if cell.data_type != 'f':
                    cell.value = status
                    cell.fill = GREEN_FILL if status == "Pass" else RED_FILL if status == "Fail" else PatternFill()
                    cell.number_format = 'General'
                    updated_count += 1
            match_log.append((fund_name, matched_name, score, status))
        else:
            match_log.append((fund_name, "No match", 0, "N/A"))

    out_bytes = io.BytesIO()
    wb.save(out_bytes)
    out_bytes.seek(0)
    return out_bytes, updated_count, match_log

# === Page Routing ===
if page == "About FidSync":
    st.header("About FidSync")
    st.markdown("""
    **FidSync** is a secure and intelligent web platform designed to streamline fiduciary tasks for financial professionals, retirement plan consultants, and compliance teams.

    Our goal is to simplify the process of reviewing fund data, documenting due diligence, and staying compliant — all in one centralized place.

    **Current capabilities include:**
    - ✅ Extracting statuses from Fund Scorecard PDFs
    - ✅ Auto-updating Excel workbooks with pass/fail results
    - ✅ Easy match verification logs for transparency

    **Coming soon:**
    - 🔍 Plan comparisons and benchmarking tools
    - 🛡️ Compliance audit checks
    - 📁 Automated audit logs

    FidSync is designed with security, accuracy, and scale in mind. Built by advisors, for advisors.
    """)

elif page == "How to Use":
    st.header("User Manual")
    st.markdown("""
    ### Step 1: Upload Your Files
    - Upload a **Fund Scorecard PDF**
    - Upload the corresponding **Excel workbook** where results should be written

    ### Step 2: Configure Settings
    - Enter the sheet name (e.g., `Current Period`)
    - Choose the **starting column letter** where statuses should be entered (e.g., `U`)
    - Enter the **starting row number** (e.g., `5`)
    - Select the **page numbers** in the PDF to scan (e.g., 20–30)

    ### Step 3: Provide Investment Names
    - Paste your list of investment options (one per line) from the Excel file

    ### Step 4: Run the Update
    - Click `Run Status Update`
    - Results will populate into the Excel file and be available for download
    - A match log will be shown and downloadable as a CSV

    **Pro Tips:**
    - Use “Dry Run” mode to preview without editing Excel
    - Check the match score column for fuzzy match accuracy
    - Excel formulas are automatically skipped (not overwritten)
    """)

elif tool == "Fund Scorecard":
    st.header("Fund Scorecard Status Tool")

    # Upload Form
    with st.form("upload_form"):
        st.subheader("Upload Files")
        pdf_file = st.file_uploader("Upload Fund Scorecard PDF", type=["pdf"])
        excel_file = st.file_uploader("Upload Excel Workbook", type=["xlsx", "xlsm"])

        st.subheader("Settings")
        col1, col2, col3 = st.columns(3)
        sheet_name = col1.text_input("Excel Sheet Name")
        status_col = col2.text_input("Starting Column Letter").strip().upper()
        start_row = col3.number_input("Starting Row Number", min_value=1)

        col4, col5 = st.columns(2)
        start_page = col4.number_input("Start Page in PDF (actual page #)", min_value=1)
        end_page = col5.number_input("End Page in PDF (actual page #)", min_value=1)

        fund_names_input = st.text_area("Investment Option Names (One Per Line)", height=200)
        dry_run = st.checkbox("Dry Run (preview match only, don't modify Excel)", value=False)

        submitted = st.form_submit_button("Run Status Update")

    # Run logic
    if submitted:
        if not pdf_file or not excel_file:
            st.warning("Please upload both PDF and Excel files.")
        elif start_page > end_page:
            st.error("Start Page must be less than or equal to End Page.")
        elif not fund_names_input.strip():
            st.warning("Please enter investment option names.")
        else:
            with st.spinner("Processing..."):
                try:
                    fund_names = [line.strip() for line in fund_names_input.strip().splitlines() if line.strip()]
                    pdf_bytes = pdf_file.read()
                    excel_bytes = excel_file.read()

                    updated_excel, count, match_log = update_excel_with_status(
                        pdf_bytes, excel_bytes, sheet_name, status_col,
                        start_row, fund_names, start_page, end_page,
                        dry_run
                    )

                    if dry_run:
                        st.info("Dry run complete. No changes were made to the Excel file.")
                    else:
                        st.success(f"Successfully updated {count} row(s).")
                        b64 = base64.b64encode(updated_excel.getvalue()).decode()
                        link = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="Updated_Investment_Status.xlsx">📥 Download Updated Excel</a>'
                        st.markdown(link, unsafe_allow_html=True)

                    df_log = pd.DataFrame(match_log, columns=["Input Name", "Matched Name", "Match Score", "Status"])
                    st.markdown("### Match Log")
                    st.dataframe(df_log)

                    csv_buffer = io.StringIO()
                    df_log.to_csv(csv_buffer, index=False)
                    csv_b64 = base64.b64encode(csv_buffer.getvalue().encode()).decode()
                    csv_link = f'<a href="data:file/csv;base64,{csv_b64}" download="match_log.csv">🧾 Download Match Log CSV</a>'
                    st.markdown(csv_link, unsafe_allow_html=True)

                    del pdf_bytes, excel_bytes, updated_excel, df_log
                    gc.collect()

                except Exception as e:
                    st.error("Something went wrong. Please check your inputs.")
                    st.exception(e)
