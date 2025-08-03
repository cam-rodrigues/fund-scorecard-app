import re
import streamlit as st
import pdfplumber
from calendar import month_name
import pandas as pd
from rapidfuzz import fuzz
from pptx import Presentation
from pptx.util import Inches, Pt
from pptx.dml.color import RGBColor
from pptx.enum.text import PP_ALIGN, MSO_VERTICAL_ANCHOR
from io import BytesIO
import yfinance as yf

#───Performance Table──────────────────────────────────────────────────────────────────

def extract_performance_table(pdf, performance_page, fund_names, end_page=None):
    import re
    from rapidfuzz import fuzz

    # Decide where to stop in the PDF
    end = end_page if end_page is not None else len(pdf.pages) + 1

    # 1. Get all lines from the section
    lines = []
    for pnum in range(performance_page - 1, end - 1):
        txt = pdf.pages[pnum].extract_text() or ""
        lines += [ln.strip() for ln in txt.splitlines() if ln.strip()]

    # 2. Prepare regex
    num_rx = re.compile(r"\(?-?\d+\.\d+%?\)?")

    # 3. For each fund, try to pull numbers
    perf_data = []
    for name in fund_names:
        item = {"Fund Scorecard Name": name}
        tk = ""  # You’ll fill in ticker later from st.session_state["tickers"] or similar
        # a) Exact-ticker match: you’ll want to match this if you have tickers already
        idx = next(
            (i for i, ln in enumerate(lines)
             if name in ln),
            None
        )
        # b) Fuzzy-name fallback if not found
        if idx is None:
            scores = [(i, fuzz.token_sort_ratio(name.lower(), ln.lower()))
                      for i, ln in enumerate(lines)]
            best_i, best_score = max(scores, key=lambda x: x[1])
            if best_score > 60:
                idx = best_i
            else:
                continue  # Can't find a match

        # c) Pull fund numbers from line above (and two above if needed)
        raw = num_rx.findall(lines[idx - 1]) if idx >= 1 else []
        if len(raw) < 8 and idx >= 2:
            raw = num_rx.findall(lines[idx - 2]) + raw
        clean = [n.strip("()%").rstrip("%") for n in raw]
        clean += [None] * (8 - len(clean))  # pad

        # d) Map to columns
        item["QTD"] = clean[0]
        item["1Yr"] = clean[2]
        item["3Yr"] = clean[3]
        item["5Yr"] = clean[4]
        item["10Yr"] = clean[5]
        item["Net Expense Ratio"] = clean[-2]

        # e) Pull benchmark QTD, 3Yr, 5Yr from next lines
        bench_raw = []
        if idx + 1 < len(lines):
            bench_raw = num_rx.findall(lines[idx + 1])
        if len(bench_raw) < 1 and idx + 2 < len(lines):
            bench_raw = num_rx.findall(lines[idx + 2])
        bench_clean = [n.strip("()%").rstrip("%") for n in bench_raw]
        item["Bench QTD"] = bench_clean[0] if bench_clean else None
        item["Bench 3Yr"] = bench_clean[3] if len(bench_clean) > 3 else None
        item["Bench 5Yr"] = bench_clean[4] if len(bench_clean) > 4 else None

        perf_data.append(item)
    return perf_data

#───Utility──────────────────────────────────────────────────────────────────

def extract_report_date(text):
    """
    Extracts and formats the report date from a block of text.
    Returns a string like '2nd QTR, 2024' for quarter-end dates,
    or 'As of March 12, 2024' for other dates.
    """
    dates = re.findall(r'(\d{1,2})/(\d{1,2})/(20\d{2})', text or "")
    for month, day, year in dates:
        m, d = int(month), int(day)
        # Quarter-end mapping
        quarter_map = {(3,31): "1st", (6,30): "2nd", (9,30): "3rd", (12,31): "4th"}
        if (m, d) in quarter_map:
            return f"{quarter_map[(m, d)]} QTR, {year}"
        # Fallback: month-day-year as human readable
        return f"As of {month_name[m]} {d}, {year}"
    return None

#───Page 1──────────────────────────────────────────────────────────────────

def process_page1(text):
    """
    Extracts 'report_date', 'total_options', 'prepared_for', and 'prepared_by' from text.
    Populates st.session_state with those keys.
    """
    # Extract report date
    report_date = extract_report_date(text)
    if report_date:
        st.session_state['report_date'] = report_date
    else:
        st.session_state['report_date'] = None

    # Extract total options
    m = re.search(r"Total Options:\s*(\d+)", text or "")
    st.session_state['total_options'] = int(m.group(1)) if m else None

    # Extract prepared for
    m = re.search(r"Prepared For:\s*\n(.*)", text or "")
    st.session_state['prepared_for'] = m.group(1).strip() if m else None

    # Extract prepared by, default if blank or says "mpi stylus"
    m = re.search(r"Prepared By:\s*(.*)", text or "")
    pb = m.group(1).strip() if m else ""
    if not pb or "mpi stylus" in pb.lower():
        pb = "Procyon Partners, LLC"
    st.session_state['prepared_by'] = pb

#───Info Card──────────────────────────────────────────────────────────────────

def show_report_summary():
    report_date    = st.session_state.get('report_date', 'N/A')
    total_options  = st.session_state.get('total_options', 'N/A')
    prepared_for   = st.session_state.get('prepared_for', 'N/A')
    prepared_by    = st.session_state.get('prepared_by', 'N/A')

    st.markdown(f"""
        <div style="
            background: linear-gradient(120deg, #e6f0fb 80%, #c8e0f6 100%);
            color: #244369;
            border-radius: 1.5rem;
            box-shadow: 0 4px 24px rgba(44,85,130,0.11), 0 2px 8px rgba(36,67,105,0.09);
            padding: 1.2rem 2.3rem 1.2rem 2.3rem;
            margin-bottom: 2rem;
            font-size: 1.08rem;
            border: 1.2px solid #b5d0eb;">
            <div style="color:#244369;">
                <b>Report Date:</b> {report_date} <br>
                <b>Total Options:</b> {total_options} <br>
                <b>Prepared For:</b> {prepared_for} <br>
                <b>Prepared By:</b> {prepared_by}
            </div>
        </div>
    """, unsafe_allow_html=True)

#───Step 2: Table of Contents Extraction──────────────────────────────────────────────────────────

def process_toc(text):
    perf = re.search(r"Fund Performance[^\d]*(\d{1,3})", text or "")
    cy   = re.search(r"Fund Performance: Calendar Year\s+(\d{1,3})", text or "")
    r3yr = re.search(r"Risk Analysis: MPT Statistics \(3Yr\)\s+(\d{1,3})", text or "")
    r5yr = re.search(r"Risk Analysis: MPT Statistics \(5Yr\)\s+(\d{1,3})", text or "")

    sc            = re.search(r"Fund Scorecard\s+(\d{1,3})", text or "")
    sc_prop       = re.search(r"Fund Scorecard:\s*Proposed Funds\s+(\d{1,3})", text or "")

    fs            = re.search(r"Fund Factsheets\s+(\d{1,3})", text or "")
    fs_prop       = re.search(r"Fund Factsheets:\s*Proposed Funds\s+(\d{1,3})", text or "")

    perf_page     = int(perf.group(1)) if perf else None
    cy_page       = int(cy.group(1)) if cy else None
    r3yr_page     = int(r3yr.group(1)) if r3yr else None
    r5yr_page     = int(r5yr.group(1)) if r5yr else None
    sc_page       = int(sc.group(1)) if sc else None
    sc_prop_page  = int(sc_prop.group(1)) if sc_prop else None
    fs_page       = int(fs.group(1)) if fs else None
    fs_prop_page  = int(fs_prop.group(1)) if fs_prop else None

    # Store in session state for downstream use
    st.session_state['performance_page'] = perf_page
    st.session_state['calendar_year_page'] = cy_page
    st.session_state['r3yr_page'] = r3yr_page
    st.session_state['r5yr_page'] = r5yr_page
    st.session_state['scorecard_page'] = sc_page
    st.session_state['scorecard_proposed_page'] = sc_prop_page
    st.session_state['factsheets_page'] = fs_page
    st.session_state['factsheets_proposed_page'] = fs_prop_page



#───IPS Invesment Screening──────────────────────────────────────────────────────────────────
import streamlit as st
import re
import pdfplumber
import pandas as pd
import yfinance as yf

def infer_fund_type_guess(ticker):
    """Infer 'Active' or 'Passive' based on Yahoo Finance info (name and summary)."""
    try:
        if not ticker:
            return ""
        info = yf.Ticker(ticker).info
        name = (info.get("longName") or info.get("shortName") or "").lower()
        summary = (info.get("longBusinessSummary") or "").lower()
        if "index" in name or "index" in summary:
            return "Passive"
        if "track" in summary and "index" in summary:
            return "Passive"
        if "actively managed" in summary or "actively-managed" in summary:
            return "Active"
        if "outperform" in summary or "manager selects" in summary:
            return "Active"
        return ""
    except Exception:
        return ""

def extract_scorecard_blocks(pdf, scorecard_page):
    metric_labels = [
        "Manager Tenure", "Excess Performance (3Yr)", "Excess Performance (5Yr)",
        "Peer Return Rank (3Yr)", "Peer Return Rank (5Yr)", "Expense Ratio Rank",
        "Sharpe Ratio Rank (3Yr)", "Sharpe Ratio Rank (5Yr)", "R-Squared (3Yr)",
        "R-Squared (5Yr)", "Sortino Ratio Rank (3Yr)", "Sortino Ratio Rank (5Yr)",
        "Tracking Error Rank (3Yr)", "Tracking Error Rank (5Yr)"
    ]
    pages, fund_blocks, fund_name, metrics = [], [], None, []
    for p in pdf.pages[scorecard_page-1:]:
        pages.append(p.extract_text() or "")
    lines = "\n".join(pages).splitlines()
    for line in lines:
        if not any(metric in line for metric in metric_labels) and line.strip():
            if fund_name and metrics:
                fund_blocks.append({"Fund Name": fund_name, "Metrics": metrics})
            fund_name = re.sub(r"Fund (Meets Watchlist Criteria|has been placed on watchlist for not meeting .* out of 14 criteria)", "", line.strip()).strip()
            metrics = []
        for metric in metric_labels:
            if metric in line:
                m = re.match(r"^(.*?)\s+(Pass|Review|Fail)\s*(.*)", line.strip())
                if m:
                    metric_name, status, info = m.groups()
                    metrics.append({"Metric": metric_name, "Status": status, "Info": info.strip()})
    if fund_name and metrics:
        fund_blocks.append({"Fund Name": fund_name, "Metrics": metrics})
    return fund_blocks

def extract_fund_tickers(pdf, performance_page, fund_names, factsheets_page=None):
    import re
    from rapidfuzz import fuzz

    def normalize_name(name):
        # strip the watchlist suffix and punctuation, lowercase
        cleaned = re.sub(
            r"has been placed on watchlist for not meeting .*? criteria",
            "",
            name,
            flags=re.IGNORECASE
        )
        cleaned = re.sub(r"[^A-Za-z0-9 ]+", "", cleaned)  # remove punctuation
        return cleaned.strip().lower()

    end_page = factsheets_page - 1 if factsheets_page else len(pdf.pages)
    all_lines = []
    for p in pdf.pages[performance_page - 1:end_page]:
        txt = p.extract_text() or ""
        all_lines.extend([ln.strip() for ln in txt.splitlines() if ln.strip()])

    # Step 1: Collect candidate (raw_name, ticker) pairs from lines with uppercase tickers length 2-5
    candidate_pairs = []  # list of tuples (normalized_raw_name, ticker, raw_name_original)
    ticker_rx = re.compile(r"\b([A-Z]{2,5})\b")
    for ln in all_lines:
        matches = ticker_rx.findall(ln)
        if not matches:
            continue
        for ticker in set(matches):
            parts = ln.rsplit(ticker, 1)
            if len(parts) >= 1:
                raw_name = parts[0].strip()
                if not raw_name:
                    continue
                norm_raw = normalize_name(raw_name)
                if not norm_raw:
                    continue
                candidate_pairs.append((norm_raw, ticker, raw_name))

    # Step 2: For each fund_name, find best candidate fuzzy match (one-to-one)
    assigned = {}
    used_tickers = set()

    # Precompute normalized expected names
    norm_expected_list = [
        (name, normalize_name(name))
        for name in fund_names
    ]

    for fund_name, norm_expected in norm_expected_list:
        best = (None, None, 0)  # (ticker, raw_name_original, score)
        for norm_raw, ticker, raw_name in candidate_pairs:
            if ticker in used_tickers:
                continue
            score = fuzz.token_sort_ratio(norm_expected, norm_raw)
            if score > best[2]:
                best = (ticker, raw_name, score)
        if best[2] >= 70:  # threshold, adjust if needed
            assigned[fund_name] = best[0]
            used_tickers.add(best[0])
        else:
            assigned[fund_name] = ""

    # Step 3: Fallback for unmatched names: looser heuristic and line-based fallback
    for name in fund_names:
        if assigned.get(name):
            continue
        norm_expected = normalize_name(name)
        for norm_raw, ticker, raw_name in candidate_pairs:
            if ticker in used_tickers:
                continue
            if norm_expected in norm_raw or norm_raw in norm_expected:
                assigned[name] = ticker
                used_tickers.add(ticker)
                break
        if not assigned.get(name):
            for ln in all_lines:
                if name.lower() in ln.lower():
                    m = re.search(r"\b([A-Z]{2,5})\b", ln)
                    if m:
                        assigned[name] = m.group(1)
                        break

    # Final cleanup: ensure non-None strings
    return {k: (v if v else "") for k, v in assigned.items()}

def scorecard_to_ips(fund_blocks, fund_types, tickers):
    # Maps for converting scorecard to IPS criteria (from your logic)
    metrics_order = [
        "Manager Tenure", "Excess Performance (3Yr)", "Excess Performance (5Yr)",
        "Peer Return Rank (3Yr)", "Peer Return Rank (5Yr)", "Expense Ratio Rank",
        "Sharpe Ratio Rank (3Yr)", "Sharpe Ratio Rank (5Yr)", "R-Squared (3Yr)",
        "R-Squared (5Yr)", "Sortino Ratio Rank (3Yr)", "Sortino Ratio Rank (5Yr)",
        "Tracking Error Rank (3Yr)", "Tracking Error Rank (5Yr)",
    ]
    active_map  = [0,1,3,6,10,2,4,7,11,5,None]
    passive_map = [0,8,3,6,12,9,4,7,13,5,None]
    ips_labels = [f"IPS Investment Criteria {i+1}" for i in range(11)]
    ips_results, raw_results = [], []
    for fund in fund_blocks:
        fund_name = fund["Fund Name"]
        fund_type = fund_types.get(fund_name, "Passive" if "index" in fund_name.lower() else "Active")
        metrics = fund["Metrics"]
        scorecard_status = [next((m["Status"] for m in metrics if m["Metric"] == label), None) for label in metrics_order]
        idx_map = passive_map if fund_type == "Passive" else active_map
        ips_status = [scorecard_status[m_idx] if m_idx is not None else "Pass" for m_idx in idx_map]
        review_fail = sum(1 for status in ips_status if status in ["Review","Fail"])
        watch_status = "FW" if review_fail >= 6 else "IW" if review_fail >= 5 else "NW"
        def iconify(status): return "✔" if status == "Pass" else "✗" if status in ("Review", "Fail") else ""
        row = {
            "Fund Name": fund_name,
            "Ticker": tickers.get(fund_name, ""),
            "Fund Type": fund_type,
            **{ips_labels[i]: iconify(ips_status[i]) for i in range(11)},
            "IPS Watch Status": watch_status,
        }
        ips_results.append(row)
        raw_results.append({
            "Fund Name": fund_name,
            "Ticker": tickers.get(fund_name, ""),
            "Fund Type": fund_type,
            **{ips_labels[i]: ips_status[i] for i in range(11)},
            "IPS Watch Status": watch_status,
        })
    return pd.DataFrame(ips_results), pd.DataFrame(raw_results)

def watch_status_color(val):
    if val == "FW":
        return "background-color:#f8d7da; color:#c30000; font-weight:600;"
    if val == "IW":
        return "background-color:#fff3cd; color:#B87333; font-weight:600;"
    if val == "NW":
        return "background-color:#d6f5df; color:#217a3e; font-weight:600;"
    return ""

def step3_5_6_scorecard_and_ips(pdf, scorecard_page, performance_page, factsheets_page, total_options):
    import pandas as pd
    import streamlit as st

    # --- 1. Extract scorecard blocks ---
    fund_blocks = extract_scorecard_blocks(pdf, scorecard_page)
    fund_names = [fund["Fund Name"] for fund in fund_blocks]
    if not fund_blocks:
        st.error("Could not extract fund scorecard blocks. Check the PDF and page number.")
        return

    # --- 2. Extract tickers ---
    tickers = extract_fund_tickers(pdf, performance_page, fund_names, factsheets_page)

    # --- Prepare inferred fund type guesses/defaults ---
    inferred_guesses = []
    for name in fund_names:
        guess = ""
        if tickers.get(name):
            guess = infer_fund_type_guess(tickers.get(name, "")) or ""
        inferred_guesses.append("Passive" if guess.lower() == "passive" else ("Passive" if "index" in name.lower() else "Active"))

    # Build base df for editor
    df_types_base = pd.DataFrame({
        "Fund Name":       fund_names,
        "Ticker":          [tickers.get(n, "") for n in fund_names],
        "Inferred Type":   inferred_guesses,
        "Fund Type":       inferred_guesses,  # starts same as inferred
    })

    # --- Toggleable editor display ---
    if "show_edit_fund_type" not in st.session_state:
        st.session_state["show_edit_fund_type"] = False
    toggle_label = "Hide Fund Type Editor" if st.session_state["show_edit_fund_type"] else "Edit Fund Type"
    if st.button(toggle_label, key="toggle_edit_fund_type"):
        st.session_state["show_edit_fund_type"] = not st.session_state["show_edit_fund_type"]

    # Decide fund_types mapping
    if st.session_state["show_edit_fund_type"]:
        st.markdown("### Fund Type Overrides")
        st.caption("Edit any fund type here; once you modify at least one, your edits take precedence over inferred types.")
        edited_types = st.data_editor(
            df_types_base,
            column_config={
                "Fund Type": st.column_config.SelectboxColumn("Fund Type", options=["Active", "Passive"]),
            },
            disabled=["Fund Name", "Ticker", "Inferred Type"],
            hide_index=True,
            key="data_editor_fundtype_ips",
            use_container_width=True,
        )
        # Determine whether any manual edit occurred (diff between Fund Type and Inferred Type)
        manual_override = any(
            row["Fund Type"] != row["Inferred Type"]
            for _, row in edited_types.iterrows()
        )
        if manual_override:
            fund_types = {row["Fund Name"]: row["Fund Type"] for _, row in edited_types.iterrows()}
        else:
            # none edited: use inferred
            fund_types = {row["Fund Name"]: row["Inferred Type"] for _, row in edited_types.iterrows()}
    else:
        # editor hidden: use inferred guesses
        fund_types = {name: inferred_guesses[i] for i, name in enumerate(fund_names)}

    # --- 4. IPS conversion ---
    df_icon, df_raw = scorecard_to_ips(fund_blocks, fund_types, tickers)

    # --- IPS Results ---
    st.subheader("IPS Screening Results")
    st.markdown(
        '<div style="display:flex; gap:1rem; margin-bottom:0.5rem;">'
        '<div style="padding:4px 10px; background:#d6f5df; border-radius:4px; font-weight:600;">NW: No Watch</div>'
        '<div style="padding:4px 10px; background:#fff3cd; border-radius:4px; font-weight:600;">IW: Informal Watch</div>'
        '<div style="padding:4px 10px; background:#f8d7da; border-radius:4px; font-weight:600;">FW: Formal Watch</div>'
        '</div>', unsafe_allow_html=True
    )

    if df_icon.empty:
        st.info("No IPS screening data available.")
    else:
        # Compact numbered criteria
        display_columns = {f"IPS Investment Criteria {i+1}": str(i+1) for i in range(11)}
        display_df = df_icon.rename(columns=display_columns)

        def iconify(s):
            if s == "Pass":
                return "✔"
            if s in ("Review", "Fail"):
                return "✗"
            return ""

        compact_df = display_df.copy()
        for i in range(1, 12):
            orig = f"IPS Investment Criteria {i}"
            if orig in compact_df.columns:
                compact_df[str(i)] = compact_df[orig].apply(iconify)
                compact_df.drop(columns=[orig], inplace=True)
        cols_order = ["Fund Name", "Ticker", "Fund Type"] + [str(i) for i in range(1, 12)] + ["IPS Watch Status"]
        compact_df = compact_df[[c for c in cols_order if c in compact_df.columns]]

        def watch_style(val):
            if val == "NW":
                return "background-color:#d6f5df; color:#217a3e; font-weight:600;"
            if val == "IW":
                return "background-color:#fff3cd; color:#B87333; font-weight:600;"
            if val == "FW":
                return "background-color:#f8d7da; color:#c30000; font-weight:600;"
            return ""

        styled = compact_df.style.applymap(watch_style, subset=["IPS Watch Status"])
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # --- Summary badges (bottom) ---
        def summarize_watch(df):
            counts = df["IPS Watch Status"].value_counts().to_dict()
            return {
                "No Watch": counts.get("NW", 0),
                "Informal Watch": counts.get("IW", 0),
                "Formal Watch": counts.get("FW", 0),
            }

        summary = summarize_watch(df_icon)
        st.markdown("---")
        st.markdown("### Watch Summary")
        b1, b2, b3 = st.columns(3, gap="small")
        with b1:
            st.metric("No Watch", summary["No Watch"])
        with b2:
            st.metric("Informal Watch", summary["Informal Watch"])
        with b3:
            st.metric("Formal Watch", summary["Formal Watch"])

    # --- Persist state downstream ---
    st.session_state["fund_blocks"] = fund_blocks
    st.session_state["fund_types"] = fund_types
    st.session_state["fund_tickers"] = tickers
    st.session_state["ips_icon_table"] = df_icon
    st.session_state["ips_raw_table"] = df_raw

    # --- Performance extraction ---
    perf_data = extract_performance_table(
        pdf,
        performance_page,
        fund_names,
        factsheets_page
    )
    for itm in perf_data:
        itm["Ticker"] = tickers.get(itm["Fund Scorecard Name"], "")

    st.session_state["fund_performance_data"] = perf_data
    st.session_state["tickers"] = tickers  # legacy compatibility


#───Step 6:Factsheets Pages──────────────────────────────────────────────────────────────────

def step6_process_factsheets(pdf, fund_names, suppress_output=True):
    # If you ever want UI, set suppress_output=False when calling

    factsheet_start = st.session_state.get("factsheets_page")
    total_declared = st.session_state.get("total_options")
    performance_data = [
        {"Fund Scorecard Name": name, "Ticker": ticker}
        for name, ticker in st.session_state.get("tickers", {}).items()
    ]

    if not factsheet_start:
        if not suppress_output:
            st.error("❌ 'Fund Factsheets' page number not found in TOC.")
        return

    matched_factsheets = []
    for i in range(factsheet_start - 1, len(pdf.pages)):
        page = pdf.pages[i]
        words = page.extract_words(use_text_flow=True)
        header_words = [w['text'] for w in words if w['top'] < 100]
        first_line = " ".join(header_words).strip()

        if not first_line or "Benchmark:" not in first_line or "Expense Ratio:" not in first_line:
            continue

        ticker_match = re.search(r"\b([A-Z]{5})\b", first_line)
        ticker = ticker_match.group(1) if ticker_match else ""
        fund_name_raw = first_line.split(ticker)[0].strip() if ticker else first_line

        best_score = 0
        matched_name = matched_ticker = ""
        for item in performance_data:
            ref = f"{item['Fund Scorecard Name']} {item['Ticker']}".strip()
            score = fuzz.token_sort_ratio(f"{fund_name_raw} {ticker}".lower(), ref.lower())
            if score > best_score:
                best_score, matched_name, matched_ticker = score, item['Fund Scorecard Name'], item['Ticker']

        def extract_field(label, text, stop=None):
            try:
                start = text.index(label) + len(label)
                rest = text[start:]
                if stop and stop in rest:
                    return rest[:rest.index(stop)].strip()
                return rest.split()[0]
            except Exception:
                return ""

        benchmark = extract_field("Benchmark:", first_line, "Category:")
        category  = extract_field("Category:", first_line, "Net Assets:")
        net_assets= extract_field("Net Assets:", first_line, "Manager Name:")
        manager   = extract_field("Manager Name:", first_line, "Avg. Market Cap:")
        avg_cap   = extract_field("Avg. Market Cap:", first_line, "Expense Ratio:")
        expense   = extract_field("Expense Ratio:", first_line)

        matched_factsheets.append({
            "Page #": i + 1,
            "Parsed Fund Name": fund_name_raw,
            "Parsed Ticker": ticker,
            "Matched Fund Name": matched_name,
            "Matched Ticker": matched_ticker,
            "Benchmark": benchmark,
            "Category": category,
            "Net Assets": net_assets,
            "Manager Name": manager,
            "Avg. Market Cap": avg_cap,
            "Expense Ratio": expense,
            "Match Score": best_score,
            "Matched": "✅" if best_score > 20 else "❌"
        })

    df_facts = pd.DataFrame(matched_factsheets)
    st.session_state['fund_factsheets_data'] = matched_factsheets

    # Hide UI output unless suppress_output is False
    if not suppress_output:
        display_df = df_facts[[
            "Matched Fund Name", "Matched Ticker", "Benchmark", "Category",
            "Net Assets", "Manager Name", "Avg. Market Cap", "Expense Ratio", "Matched"
        ]].rename(columns={"Matched Fund Name": "Fund Name", "Matched Ticker": "Ticker"})

        st.dataframe(display_df, use_container_width=True)

        matched_count = sum(1 for r in matched_factsheets if r["Matched"] == "✅")
        st.write(f"Matched {matched_count} of {len(matched_factsheets)} factsheet pages.")
        if matched_count == total_declared:
            st.success(f"All {matched_count} funds matched the declared Total Options from Page 1.")
        else:
            st.error(f"Mismatch: Page 1 declared {total_declared}, but only matched {matched_count}.")

#───Step 7: QTD / 1Yr / 3Yr / 5Yr / 10Yr / Net Expense Ratio & Bench QTD──────────────────────────────────────────────────────────────────

def step7_extract_returns(pdf):
    import re
    import pandas as pd
    import streamlit as st
    from rapidfuzz import fuzz

    # 1) Where to scan
    perf_page = st.session_state.get("performance_page")
    end_page  = st.session_state.get("calendar_year_page") or (len(pdf.pages) + 1)
    perf_data = st.session_state.get("fund_performance_data", [])
    if perf_page is None or not perf_data:
        st.error("❌ Run Step 5 first to populate performance data.")
        return

    # 2) Prep output slots
    fields = [
        "QTD", "1Yr", "3Yr", "5Yr", "10Yr", "Net Expense Ratio",
        "Bench QTD", "Bench 1Yr", "Bench 3Yr", "Bench 5Yr", "Bench 10Yr"
    ]
    for itm in perf_data:
        for f in fields:
            itm.setdefault(f, None)

    # 3) Gather every nonblank line in the Performance section
    lines = []
    for pnum in range(perf_page - 1, end_page - 1):
        txt = pdf.pages[pnum].extract_text() or ""
        lines += [ln.strip() for ln in txt.splitlines() if ln.strip()]

    # 4) Regex to pull decimal tokens (with optional % and parentheses)
    num_rx = re.compile(r"\(?-?\d+\.\d+%?\)?")

    matched = 0
    for item in perf_data:
        name = item["Fund Scorecard Name"]
        tk   = item["Ticker"].upper().strip()

        # a) Exact-ticker match
        idx = next(
            (i for i, ln in enumerate(lines)
             if re.search(rf"\b{re.escape(tk)}\b", ln)),
            None
        )
        # b) Fuzzy-name fallback
        if idx is None:
            scores = [(i, fuzz.token_sort_ratio(name.lower(), ln.lower()))
                      for i, ln in enumerate(lines)]
            best_i, best_score = max(scores, key=lambda x: x[1])
            if best_score > 60:
                idx = best_i
            else:
                st.warning(f"⚠️ {name} ({tk}): no match found.")
                continue

        # c) Pull fund numbers from line above (and two above if needed)
        raw = num_rx.findall(lines[idx - 1]) if idx >= 1 else []
        if len(raw) < 8 and idx >= 2:
            raw = num_rx.findall(lines[idx - 2]) + raw
        clean = [n.strip("()%").rstrip("%") for n in raw]
        if len(clean) < 8:
            clean += [None] * (8 - len(clean))

        # d) Map fund returns & net expense
        item["QTD"]               = clean[0]
        item["1Yr"]               = clean[2]
        item["3Yr"]               = clean[3]
        item["5Yr"]               = clean[4]
        item["10Yr"]              = clean[5]
        item["Net Expense Ratio"] = clean[-2]

        # e) Pull benchmark QTD, 1Yr, 3Yr, 5Yr, and 10Yr from the very next line(s)
        bench_raw = []
        if idx + 1 < len(lines):
            bench_raw = num_rx.findall(lines[idx + 1])
        if len(bench_raw) < 5 and idx + 2 < len(lines):
            bench_raw = num_rx.findall(lines[idx + 2])
        bench_clean = [n.strip("()%").rstrip("%") for n in bench_raw]

        item["Bench QTD"]  = bench_clean[0] if len(bench_clean) > 0 else None
        item["Bench 1Yr"] = bench_clean[1] if len(bench_clean) > 1 else None
        item["Bench 3Yr"] = bench_clean[3] if len(bench_clean) > 3 else None
        item["Bench 5Yr"] = bench_clean[4] if len(bench_clean) > 4 else None
        item["Bench 10Yr"] = bench_clean[5] if len(bench_clean) > 5 else None

        matched += 1

    # 5) Save & display
    st.session_state["fund_performance_data"] = perf_data
    df = pd.DataFrame(perf_data)
    # Hide the per-fund warnings and overall success message, but still alert if something is off
    
    expected_count = len(perf_data)
    if matched < expected_count:
        st.error(f"❌ Only matched {matched} of {expected_count} funds with return data. Check your PDF or extraction logic.")
    # (Do NOT display per-fund warnings or success)

    st.dataframe(
        df[["Fund Scorecard Name", "Ticker"] + fields],
        use_container_width=True
    )


#───Step 8 Calendar Year Returns (funds + benchmarks──────────────────────────────────────────────────────────────────

def step8_calendar_returns(pdf):
    import re, streamlit as st, pandas as pd

    # 1) Figure out section bounds
    cy_page  = st.session_state.get("calendar_year_page")
    end_page = st.session_state.get("r3yr_page", len(pdf.pages) + 1)
    if cy_page is None:
        st.error("❌ 'Fund Performance: Calendar Year' not found in TOC.")
        return

    # 2) Pull every line from that section
    all_lines = []
    for p in pdf.pages[cy_page-1 : end_page-1]:
        all_lines.extend((p.extract_text() or "").splitlines())

    # 3) Identify header & years
    header = next((ln for ln in all_lines if "Ticker" in ln and re.search(r"20\d{2}", ln)), None)
    if not header:
        st.error("❌ Couldn’t find header row with 'Ticker' + year.")
        return
    years = re.findall(r"\b20\d{2}\b", header)
    num_rx = re.compile(r"-?\d+\.\d+%?")

    # — A) Funds themselves —
    fund_map     = st.session_state.get("tickers", {})
    fund_records = []
    for name, tk in fund_map.items():
        ticker = (tk or "").upper()
        idx    = next((i for i, ln in enumerate(all_lines) if ticker in ln.split()), None)
        raw    = num_rx.findall(all_lines[idx-1]) if idx not in (None, 0) else []
        vals   = raw[:len(years)] + [None] * (len(years) - len(raw))
        rec    = {"Name": name, "Ticker": ticker}
        rec.update({years[i]: vals[i] for i in range(len(years))})
        fund_records.append(rec)

    df_fund = pd.DataFrame(fund_records)
    if not df_fund.empty:
        st.markdown("**Fund Calendar‑Year Returns**")
        st.dataframe(df_fund[["Name", "Ticker"] + years], use_container_width=True)
        st.session_state["step8_returns"] = fund_records

    # — B) Benchmarks matched back to each fund’s ticker —
    facts         = st.session_state.get("fund_factsheets_data", [])
    bench_records = []
    for f in facts:
        bench_name = f.get("Benchmark", "").strip()
        fund_tkr   = f.get("Matched Ticker", "")
        if not bench_name:
            continue

        # find the first line containing the benchmark name
        idx = next((i for i, ln in enumerate(all_lines) if bench_name in ln), None)
        if idx is None:
            continue
        raw  = num_rx.findall(all_lines[idx])
        vals = raw[:len(years)] + [None] * (len(years) - len(raw))
        rec  = {"Name": bench_name, "Ticker": fund_tkr}
        rec.update({years[i]: vals[i] for i in range(len(years))})
        bench_records.append(rec)

    df_bench = pd.DataFrame(bench_records)
    if not df_bench.empty:
        st.markdown("**Benchmark Calendar‑Year Returns**")
        st.dataframe(df_bench[["Name", "Ticker"] + years], use_container_width=True)
        st.session_state["benchmark_calendar_year_returns"] = bench_records
    else:
        st.warning("No benchmark returns extracted.")

#───Step 9: 3‑Yr Risk Analysis – Match & Extract MPT Stats (hidden matching)──────────────────────────────────────────────────────────────────

def step9_risk_analysis_3yr(pdf):
    import re, streamlit as st, pandas as pd
    from rapidfuzz import fuzz

    # 1) Get your fund→ticker map
    fund_map = st.session_state.get("tickers", {})
    if not fund_map:
        st.error("❌ No ticker mapping found. Run Step 5 first.")
        return

    # 2) Locate the “Risk Analysis: MPT Statistics (3Yr)” page
    start_page = st.session_state.get("r3yr_page")
    if not start_page:
        st.error("❌ ‘Risk Analysis: MPT Statistics (3Yr)’ page not found; run Step 2 first.")
        return

    # 3) Scan forward until you’ve seen each ticker (no display)
    locs = {}
    for pnum in range(start_page, len(pdf.pages) + 1):
        lines = (pdf.pages[pnum-1].extract_text() or "").splitlines()
        for li, ln in enumerate(lines):
            tokens = ln.split()
            for fname, tk in fund_map.items():
                if fname in locs: 
                    continue
                if tk.upper() in tokens:
                    locs[fname] = {"page": pnum, "line": li}
        if len(locs) == len(fund_map):
            break

    # 4) Extract the first four numeric MPT stats from that same line
    num_rx = re.compile(r"-?\d+\.\d+")
    results = []
    for name, info in locs.items():
        page = pdf.pages[info["page"]-1]
        lines = (page.extract_text() or "").splitlines()
        line = lines[info["line"]]
        nums = num_rx.findall(line)
        nums += [None] * (4 - len(nums))
        alpha, beta, up, down = nums[:4]
        results.append({
            "Fund Name":               name,
            "Ticker":                  fund_map[name].upper(),
            "3 Year Alpha":            alpha,
            "3 Year Beta":             beta,
            "3 Year Upside Capture":   up,
            "3 Year Downside Capture": down
        })

    # 5) Display final table only
    st.session_state["step9_mpt_stats"] = results

#───Step 10: Risk Analysis (5Yr) – Match & Extract MPT Statistics──────────────────────────────────────────────────────────────────

def step10_risk_analysis_5yr(pdf):
    import re, streamlit as st, pandas as pd

    # 1) Your fund→ticker map from Step 5
    fund_map = st.session_state.get("tickers", {})
    if not fund_map:
        st.error("❌ No ticker mapping found. Run Step 5 first.")
        return

    # 2) Locate the “Risk Analysis: MPT Statistics (5Yr)” section
    section_page = next(
        (i for i, pg in enumerate(pdf.pages, start=1)
         if "Risk Analysis: MPT Statistics (5Yr)" in (pg.extract_text() or "")),
        None
    )
    if section_page is None:
        st.error("❌ Could not find ‘Risk Analysis: MPT Statistics (5Yr)’ section.")
        return

    # 3) Under‑the‑hood: scan pages until each ticker is located
    locs = {}
    total = len(fund_map)
    for pnum in range(section_page, len(pdf.pages) + 1):
        lines = (pdf.pages[pnum-1].extract_text() or "").splitlines()
        for li, ln in enumerate(lines):
            tokens = ln.split()
            for name, tk in fund_map.items():
                if name in locs:
                    continue
                if tk.upper() in tokens:
                    locs[name] = {"page": pnum, "line": li}
        if len(locs) == total:
            break

    # 4) Wrap‑aware extraction of the first four floats after each ticker line
    num_rx = re.compile(r"-?\d+\.\d+")
    results = []
    for name, tk in fund_map.items():
        info = locs.get(name)
        vals = [None] * 4
        if info:
            page = pdf.pages[info["page"] - 1]
            text_lines = (page.extract_text() or "").splitlines()
            idx = info["line"]
            nums = []
            # look on the line of the ticker and up to the next 2 lines
            for j in range(idx, min(idx + 3, len(text_lines))):
                nums += num_rx.findall(text_lines[j])
                if len(nums) >= 4:
                    break
            nums += [None] * (4 - len(nums))
            vals = nums[:4]
        else:
            st.warning(f"⚠️ {name} ({tk.upper()}): not found after page {section_page}.")

        alpha5, beta5, up5, down5 = vals
        results.append({
            "Fund Name":               name,
            "Ticker":                  tk.upper(),
            "5 Year Alpha":            alpha5,
            "5 Year Beta":             beta5,
            "5 Year Upside Capture":   up5,
            "5 Year Downside Capture": down5,
        })

    # 5) Save & display only the consolidated table
    st.session_state["step10_mpt_stats"] = results

#───Step 11: Combined MPT Statistics Summary──────────────────────────────────────────────────────────────────

def step11_create_summary(pdf=None):
    import pandas as pd
    import streamlit as st

    # 1) Load your 3‑Yr and 5‑Yr stats from session state
    mpt3 = st.session_state.get("step9_mpt_stats", [])
    mpt5 = st.session_state.get("step10_mpt_stats", [])
    if not mpt3 or not mpt5:
        st.error("❌ Missing MPT stats. Run Steps 9 & 10 first.")
        return

    # 2) Build DataFrames
    df3 = pd.DataFrame(mpt3)  # contains "3 Year Alpha", "3 Year Beta", etc.
    df5 = pd.DataFrame(mpt5)  # contains "5 Year Alpha", "5 Year Beta", etc.

    # 3) Merge on Fund Name & Ticker
    df = pd.merge(
        df3,
        df5,
        on=["Fund Name", "Ticker"],
        how="outer",
        suffixes=("_3yr", "_5yr")
    )

    # 4) Build the Investment Manager column
    df.insert(0, "Investment Manager", df["Fund Name"] + " (" + df["Ticker"] + ")")

    # 5) Select & order the columns
    df = df[[
        "Investment Manager",
        "3 Year Alpha",
        "5 Year Alpha",
        "3 Year Beta",
        "5 Year Beta",
        "3 Year Upside Capture",
        "3 Year Downside Capture",
        "5 Year Upside Capture",
        "5 Year Downside Capture"
    ]]

    # 6) Display
    st.session_state["step11_summary"] = df.to_dict("records")
    st.dataframe(df)

#───Step 12: Extract “FUND FACTS” & Its Table Details in One Go──────────────────────────────────────────────────────────────────

def step12_process_fund_facts(pdf):
    import re
    import streamlit as st
    import pandas as pd

    fs_start   = st.session_state.get("factsheets_page")
    factsheets = st.session_state.get("fund_factsheets_data", [])
    if not fs_start or not factsheets:
        st.error("❌ Run Step 6 first to populate your factsheet pages.")
        return

    # map factsheet pages to fund name & ticker
    page_map = {
        f["Page #"]: (f["Matched Fund Name"], f["Matched Ticker"])
        for f in factsheets
    }

    # the exact labels and the order you want them in the table
    labels = [
        "Manager Tenure Yrs.",
        "Expense Ratio",
        "Expense Ratio Rank",
        "Total Number of Holdings",
        "Turnover Ratio"
    ]

    records = []
    # scan each factsheet page
    for pnum in range(fs_start, len(pdf.pages) + 1):
        if pnum not in page_map:
            continue
        fund_name, ticker = page_map[pnum]
        lines = pdf.pages[pnum-1].extract_text().splitlines()

        for idx, line in enumerate(lines):
            if line.lstrip().upper().startswith("FUND FACTS"):
                # grab the next 8 lines (should contain your 5 labels)
                snippet = lines[idx+1 : idx+1+8]
                rec = {"Fund Name": fund_name, "Ticker": ticker}
                for lab in labels:
                    val = None
                    for ln in snippet:
                        norm = " ".join(ln.strip().split())
                        if norm.startswith(lab):
                            rest = norm[len(lab):].strip(" :\t")
                            m = re.match(r"(-?\d+\.\d+)", rest)
                            val = m.group(1) if m else (rest.split()[0] if rest else None)
                            break
                    rec[lab] = val
                records.append(rec)
                break  # move on to the next page once Fund Facts is processed

    if not records:
        st.warning("No Fund Facts tables found.")
        return

    # save & show
    st.session_state["step12_fund_facts_table"] = records

#───Step 13: Extract Risk‑Adjusted Returns Metrics──────────────────────────────────────────────────────────────────

def step13_process_risk_adjusted_returns(pdf):
    import re
    import streamlit as st
    import pandas as pd

    fs_start   = st.session_state.get("factsheets_page")
    factsheets = st.session_state.get("fund_factsheets_data", [])
    if not fs_start or not factsheets:
        st.error("❌ Run Step 6 first to populate your factsheet pages.")
        return

    # map factsheet pages to fund name & ticker
    page_map = {
        f["Page #"]: (f["Matched Fund Name"], f["Matched Ticker"])
        for f in factsheets
    }

    # which metrics to pull
    metrics = ["Sharpe Ratio", "Information Ratio", "Sortino Ratio"]
    num_rx  = re.compile(r"-?\d+\.\d+")

    records = []
    for pnum in range(fs_start, len(pdf.pages) + 1):
        if pnum not in page_map:
            continue
        fund_name, ticker = page_map[pnum]
        lines = (pdf.pages[pnum-1].extract_text() or "").splitlines()

        # find the heading
        for idx, line in enumerate(lines):
            norm = " ".join(line.strip().split()).upper()
            if norm.startswith("RISK-ADJUSTED RETURNS"):
                snippet = lines[idx+1 : idx+1+6]  # grab next few lines
                rec = {"Fund Name": fund_name, "Ticker": ticker}

                for metric in metrics:
                    # find the snippet line for this metric
                    text_line = next(
                        ( " ".join(ln.strip().split())
                          for ln in snippet
                          if ln.strip().upper().startswith(metric.upper()) ),
                        None
                    ) or ""
                    # extract up to 4 numbers
                    nums = num_rx.findall(text_line)
                    nums += [None] * (4 - len(nums))

                    # assign into rec
                    rec[f"{metric} 1Yr"]  = nums[0]
                    rec[f"{metric} 3Yr"]  = nums[1]
                    rec[f"{metric} 5Yr"]  = nums[2]
                    rec[f"{metric} 10Yr"] = nums[3]

                records.append(rec)
                break  # done with this page

    if not records:
        st.warning("No 'RISK‑ADJUSTED RETURNS' tables found.")
        return

    # save & show
    st.session_state["step13_risk_adjusted_table"] = records
    df = pd.DataFrame(records)
    st.dataframe(df, use_container_width=True)

#───Step 14: Peer Risk-Adjusted Return Rank──────────────────────────────────────────────────────────────────

def step14_extract_peer_risk_adjusted_return_rank(pdf):
    import re
    import streamlit as st
    import pandas as pd

    st.write("Peer Rank")

    factsheets = st.session_state.get("fund_factsheets_data", [])
    if not factsheets:
        st.error("❌ Run Step 6 first to populate your factsheet pages.")
        return

    page_map = {
        f["Page #"]:(f["Matched Fund Name"], f["Matched Ticker"])
        for f in factsheets
    }

    metrics = ["Sharpe Ratio", "Information Ratio", "Sortino Ratio"]
    records = []

    for pnum, (fund, ticker) in page_map.items():
        page = pdf.pages[pnum-1]
        text = page.extract_text() or ""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

        # 1) locate the Risk-Adjusted Returns header
        try:
            risk_idx = next(i for i, ln in enumerate(lines)
                            if "RISK-ADJUSTED RETURNS" in ln.upper())
        except StopIteration:
            st.warning(f"⚠️ {fund} ({ticker}): Risk-Adjusted header not found.")
            continue

        # 2) find all “1 Yr 3 Yrs 5 Yrs 10 Yrs” lines after that
        header_idxs = [i for i, ln in enumerate(lines)
                       if re.match(r"1\s*Yr", ln)]
        peer_header_idxs = [i for i in header_idxs if i > risk_idx]

        if not peer_header_idxs:
            st.warning(f"⚠️ {fund} ({ticker}): peer header not found.")
            continue

        # take the *second* header occurrence (first is Risk-Adjusted, next is Peer)
        peer_hdr = peer_header_idxs[0] if len(peer_header_idxs)==1 else peer_header_idxs[1]

        rec = {"Fund Name": fund, "Ticker": ticker}

        # 3) read the three lines immediately below that header
        for offset, metric in enumerate(metrics, start=1):
            if peer_hdr + offset < len(lines):
                parts = lines[peer_hdr + offset].split()
                # parts[0:2] = metric name words, parts[2:6] = the four integer ranks
                vals = parts[2:6] if len(parts) >= 6 else []
            else:
                vals = []

            # fill into record (pad with None if missing)
            for idx, period in enumerate(["1Yr","3Yr","5Yr","10Yr"]):
                rec[f"{metric} {period}"] = vals[idx] if idx < len(vals) else None

            if len(vals) < 4:
                st.warning(f"⚠️ {fund} ({ticker}): only {len(vals)} peer values found for '{metric}'.")

        records.append(rec)

    if not records:
        st.warning("❌ No Peer Risk-Adjusted Return Rank data extracted.")
        return

    df = pd.DataFrame(records)
    st.session_state["step14_peer_rank_table"] = records
    st.dataframe(df, use_container_width=True)

#───Step 14.5: IPS Fail Table──────────────────────────────────────────────────────────────────

def step14_5_ips_fail_table():
    import streamlit as st
    import pandas as pd

    df = st.session_state.get("ips_icon_table")
    if df is None or df.empty:
        return

    fail_df = df[df["IPS Watch Status"].isin(["FW", "IW"])][["Fund Name", "IPS Watch Status"]]
    if fail_df.empty:
        return

    table_html = fail_df.rename(columns={
        "Fund Name": "Fund",
        "IPS Watch Status": "Watch Status"
    }).to_html(index=False, border=0, justify="center", classes="ips-fail-table")

    st.markdown(f"""
    <div style='
        background: linear-gradient(120deg, #e6f0fb 85%, #c8e0f6 100%);
        color: #23395d;
        border-radius: 1.3rem;
        box-shadow: 0 2px 14px rgba(44,85,130,0.08), 0 1px 4px rgba(36,67,105,0.07);
        padding: 1.6rem 2.0rem 1.6rem 2.0rem;
        max-width: 650px;
        margin: 1.4rem auto 1.2rem auto;
        border: 1.5px solid #b5d0eb;'>
        <div style='font-weight:700; color:#23395d; font-size:1.15rem; margin-bottom:0.5rem; letter-spacing:-0.5px;'>
            Funds on Watch
        </div>
        <div style='font-size:1rem; margin-bottom:1rem; color:#23395d;'>
            The following funds failed five or more IPS criteria and are currently on watch.
        </div>
        {table_html}
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <style>
    .ips-fail-table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 0.7em;
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif;
    }
    .ips-fail-table th, .ips-fail-table td {
        border: none;
        padding: 0.48em 1.1em;
        text-align: left;
        font-size: 1.07em;
    }
    .ips-fail-table th {
        background: #244369;
        color: #fff;
        font-weight: 700;
        letter-spacing: 0.01em;
    }
    .ips-fail-table td {
        color: #244369;
    }
    .ips-fail-table tr:nth-child(even) {background: #e6f0fb;}
    .ips-fail-table tr:nth-child(odd)  {background: #f8fafc;}
    </style>
    """, unsafe_allow_html=True)

#───Step 15 - Excel──────────────────────────────────────────────────────────────────

import os
import streamlit as st
from openpyxl import load_workbook

def step15_populate_excel_template(template_path: str = "assets/investment_metrics_template.xlsx",
                                   output_path: str = "filled_investment_metrics.xlsx"):
    """
    Step 15 minimal: on button press, write 'Prepared for' into AB4 (Current Period sheet if present,
    otherwise the active/populated sheet) and save. Does not modify anything else.
    """
    if not st.button("Populate Investment Metrics Excel (Step 15: Prepared for)"):
        return None  # wait until clicked

    if not os.path.exists(template_path):
        st.error(f"Template not found at {template_path}")
        return None

    try:
        wb = load_workbook(template_path)
    except Exception as e:
        st.error(f"Failed to open template: {e}")
        return None

    # Write "Prepared for" into AB4, preferring "Current Period" sheet
    if "Current Period" in wb.sheetnames:
        ws_target = wb["Current Period"]
    else:
        # fallback to first sheet
        ws_target = wb[wb.sheetnames[0]]

    ws_target["AB4"] = "Prepared for"

    try:
        wb.save(output_path)
    except Exception as e:
        st.error(f"Failed to save populated workbook: {e}")
        return None

    st.success(f"'Prepared for' written to AB4 and saved to {output_path}")
    return output_path



#───Main App──────────────────────────────────────────────────────────────────

def run():
    import re
    st.title("Fund Scorecard Generator")
    uploaded = st.file_uploader("Upload MPI PDF to Generate Writup PPTX", type="pdf")
    if not uploaded:
        return

    with pdfplumber.open(uploaded) as pdf:
        # Step 1
        first = pdf.pages[0].extract_text() or ""
        process_page1(first)
        show_report_summary()

        # Step 2
        with st.expander("Table of Contents", expanded=False):
            toc_text = "".join((pdf.pages[i].extract_text() or "") for i in range(min(3, len(pdf.pages))))
            process_toc(toc_text)

        # --- Combined core details grouped ---
        with st.expander("All Fund Details", expanded=True):
            # 1. IPS Investment Screening
            with st.expander("IPS Investment Screening", expanded=True):
                sp = st.session_state.get("scorecard_page")
                tot = st.session_state.get("total_options")
                pp = st.session_state.get("performance_page")
                factsheets_page = st.session_state.get("factsheets_page")
                if sp and tot is not None and pp:
                    step3_5_6_scorecard_and_ips(pdf, sp, pp, factsheets_page, tot)
                else:
                    st.error("Missing scorecard, performance page, or total options")

            # 2. Fund Factsheets
            with st.expander("Fund Factsheets", expanded=True):
                names = [b['Fund Name'] for b in st.session_state.get('fund_blocks', [])]
                step6_process_factsheets(pdf, names)
                
            # 3. Extract Fund Facts sub-headings (Step 12) so Step 15 has data
            with st.expander("Fund Facts (sub-headings)", expanded=False):
                step12_process_fund_facts(pdf)
                
            # 4. Returns (annualized + calendar)
            with st.expander("Returns", expanded=False):
                step7_extract_returns(pdf)
                step8_calendar_returns(pdf)

            # 5. MPT Statistics Summary (requires risk analyses first)
            with st.expander("MPT Statistics Summary", expanded=False):
                step9_risk_analysis_3yr(pdf)
                step10_risk_analysis_5yr(pdf)
                step11_create_summary()

            # 6. Risk-Adjusted Returns and Peer Rank
            with st.expander("Risk-Adjusted Returns", expanded=False):
                step13_process_risk_adjusted_returns(pdf)
                step14_extract_peer_risk_adjusted_return_rank(pdf)

        # Data prep for bullet points (unchanged)
        report_date = st.session_state.get("report_date", "")
        m = re.match(r"(\d)(?:st|nd|rd|th)\s+QTR,\s*(\d{4})", report_date)
        quarter = m.group(1) if m else ""
        year = m.group(2) if m else ""

        for itm in st.session_state.get("fund_performance_data", []):
            qtd = float(itm.get("QTD") or 0)
            bench_qtd = float(itm.get("Bench QTD") or 0)
            itm["Perf Direction"] = "overperformed" if qtd >= bench_qtd else "underperformed"
            itm["Quarter"] = quarter
            itm["Year"] = year
            diff_bps = round((qtd - bench_qtd) * 100, 1)
            itm["QTD_bps_diff"] = str(diff_bps)
            fund_pct = f"{qtd:.2f}%"
            bench_pct = f"{bench_qtd:.2f}%"
            itm["QTD_pct_diff"] = f"{(qtd - bench_qtd):.2f}%"
            itm["QTD_vs"] = f"{fund_pct} vs. {bench_pct}"

        if "bullet_point_templates" not in st.session_state:
            st.session_state["bullet_point_templates"] = [
                "[Fund Scorecard Name] [Perf Direction] its benchmark in Q[Quarter], "
                "[Year] by [QTD_bps_diff] bps ([QTD_vs])."
            ]

        # Step 14.5: IPS Fail Table
        step14_5_ips_fail_table()

        # Step 15: Populate template (button is inside function)
        try:
            # optional: let user type a name to show in AB4
            client_name = st.text_input("Prepared for:", value=st.session_state.get("client_name", ""))
            st.session_state["client_name"] = client_name  # persist if needed
        
            filled_path = step15_populate_excel_template(
                template_path="assets/investment_metrics_template.xlsx",
                output_path="filled_investment_metrics.xlsx"
            )
            if filled_path and os.path.exists(filled_path):
                with open(filled_path, "rb") as f:
                    data = f.read()
                st.download_button(
                    "Download populated Excel",
                    data=data,
                    file_name="filled_investment_metrics.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Step 15 failed: {e}")



if __name__ == "__main__":
    run()
