import streamlit as st
import pdfplumber
import re
from difflib import SequenceMatcher
from difflib import SequenceMatcher, get_close_matches
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from io import StringIO



def run():
    st.title("IPS Write-Up Tool")

    uploaded_file = st.file_uploader("Upload MPI PDF", type=["pdf"], key="upload_mpi")
    if not uploaded_file:
        st.stop()

    with pdfplumber.open(uploaded_file) as pdf:
        # === Step 1: Metadata ===
        st.subheader("Step 1: PDF Metadata")
        st.write("**Total Pages:**", len(pdf.pages))

#------------------------------------------------------------------------------------------------------------------

        # === Step 2: Page 1 Extraction ===
        st.subheader("Step 2: Pg 1")
        page1 = pdf.pages[0]
        text = page1.extract_text()

        # Determine Quarter
        quarter = "Unknown"
        if "3/31/20" in text:
            quarter = "Q1"
        elif "6/30/20" in text:
            quarter = "Q2"
        elif "9/30/20" in text:
            quarter = "Q3"
        elif "12/31/20" in text:
            quarter = "Q4"

        st.write("**Quarter:**", quarter)

        # Total Options
        total_match = re.search(r"Total Options:\s*(\d+)", text)
        total_options = total_match.group(1) if total_match else "Not found"
        st.write("**Total Investment Options:**", total_options)

        # Prepared For
        prepared_for_match = re.search(r"Prepared For:\s*\n(.+)", text)
        prepared_for = prepared_for_match.group(1).strip() if prepared_for_match else "Not found"
        st.write("**Prepared For:**", prepared_for)

        # Prepared By
        prepared_by_match = re.search(r"Prepared By:\s*\n(.+)", text)
        prepared_by = prepared_by_match.group(1).strip() if prepared_by_match else "Not found"
        st.write("**Prepared By:**", prepared_by)

#------------------------------------------------------------------------------------------------------------------

        # === Step 4: Pg 2 - Table of Contents ===
        st.subheader("Step 4: Pg 2 - Table of Contents")
        page2 = pdf.pages[1]
        text2 = page2.extract_text()

        fund_perf_pg = "Not found"
        fund_scorecard_pg = "Not found"

        for line in text2.split("\n"):
            if "Fund Performance: Current vs. Proposed Comparison" in line:
                match = re.search(r"(\d+)$", line)
                if match:
                    fund_perf_pg = int(match.group(1))
            if "Fund Scorecard" in line:
                match = re.search(r"(\d+)$", line)
                if match:
                    fund_scorecard_pg = int(match.group(1))

        st.write("**Fund Performance Page:**", fund_perf_pg)
        st.write("**Fund Scorecard Page:**", fund_scorecard_pg)

#------------------------------------------------------------------------------------------------------------------

        # === Step 5: Fund Scorecard Section ===
        st.subheader("Step 5: Fund Scorecard Section")

        metrics_data = []
        metrics_header = []
        fund_blocks = []

        # Text cleaning pattern
        fund_status_pattern = re.compile(
            r"\s+(Fund Meets Watchlist Criteria\.|Fund has been placed on watchlist for not meeting.+)", re.IGNORECASE)

        # Read all pages starting from Fund Scorecard page
        for i in range(fund_scorecard_pg - 1, len(pdf.pages)):
            page = pdf.pages[i]
            text = page.extract_text()
            if not text or "Fund Scorecard" not in text:
                break

            lines = text.split("\n")

            # Capture Criteria Threshold section (appears once)
            if not metrics_header:
                for j in range(len(lines)):
                    if "Criteria Threshold" in lines[j]:
                        metrics_header = lines[j+1:j+15]
                        break

            # Capture fund blocks: look for lines starting with a metric, backtrack for fund name
            for j in range(len(lines)):
                if lines[j].startswith("Manager Tenure"):
                    raw_fund_line = lines[j-1].strip()
                    # Clean extra status text from the same line
                    fund_name = fund_status_pattern.sub("", raw_fund_line).strip()
                    if "criteria threshold" in fund_name.lower():
                        continue  # skip bad block that grabbed threshold box


                    fund_metrics = []
                    for k in range(j, j+14):
                        if k >= len(lines): break
                        metric_line = lines[k]
                        match = re.match(r"(.+?)\s+(Pass|Review)\s+(.*)", metric_line)
                        if match:
                            metric_name, status, reason = match.groups()
                            fund_metrics.append({
                                "Metric": metric_name.strip(),
                                "Status": status,
                                "Reason": reason.strip()
                            })
                    fund_blocks.append({
                        "Fund Name": fund_name,
                        "Metrics": fund_metrics
                    })

        # Display criteria header
        if metrics_header:
            st.markdown("**Criteria Threshold (14 Metrics):**")
            for item in metrics_header:
                st.markdown(f"- {item}")

        # Display fund blocks
        for block in fund_blocks:
            st.markdown(f"**{block['Fund Name']}**")
            for metric in block["Metrics"]:
                st.write(f"- {metric['Metric']}: **{metric['Status']}** – {metric['Reason']}")
                
#------------------------------------------------------------------------------------------------------------------
        
        # === Step 7: Double Check ===
        st.subheader("Step 7: Double Check")

        num_extracted_funds = len(fund_blocks)
        st.write("**Investment Options Extracted:**", num_extracted_funds)

        try:
            total_expected = int(total_options)
            if num_extracted_funds == total_expected:
                st.success(f"✅ Count matches: {num_extracted_funds} funds found.")
            else:
                st.warning(f"⚠️ Count mismatch: Found {num_extracted_funds}, but expected {total_expected}.")
        except:
            st.error("❌ Unable to validate count due to unreadable 'Total Options' value.")


#------------------------------------------------------------------------------------------------------------------

        # === Step 8: IPS Investment Criteria Screening ===
        st.subheader("Step 8: IPS Investment Criteria Screening")

        ips_criteria = [
            "Manager Tenure ≥ 3 years",
            "3-Year Performance > Benchmark / +3-Year R² > 95%",
            "3-Year Performance > 50% of Peers",
            "3-Year Sharpe Ratio > 50% of Peers",
            "3-Year Sortino Ratio > 50% of Peers / +3-Year Tracking Error < 90% of Peers",
            "5-Year Performance > Benchmark / +5-Year R² > 95%",
            "5-Year Performance > 50% of Peers",
            "5-Year Sharpe Ratio > 50% of Peers",
            "5-Year Sortino Ratio > 50% of Peers / +5-Year Tracking Error < 90% of Peers",
            "Expense Ratio < 50% of Peers",
            "Investment Style aligns with fund objectives"
        ]

        st.markdown("**IPS Investment Criteria:**")
        for i, crit in enumerate(ips_criteria, 1):
            st.markdown(f"{i}. {crit}")

        # Define mappings to fund scorecard metric labels
        def map_metric_names(fund_type):
            if fund_type == "Passive":
                return [
                    "Manager Tenure",
                    "R² (3Yr)",
                    "Return Rank (3Yr)",
                    "Sharpe Ratio Rank (3Yr)",
                    "Tracking Error Rank (3Yr)",
                    "R² (5Yr)",
                    "Return Rank (5Yr)",
                    "Sharpe Ratio Rank (5Yr)",
                    "Tracking Error Rank (5Yr)",
                    "Expense Ratio Rank",
                    "Investment Style"
                ]
            else:  # Active
                return [
                    "Manager Tenure",
                    "Excess Performance (3Yr)",
                    "Return Rank (3Yr)",
                    "Sharpe Ratio Rank (3Yr)",
                    "Sortino Ratio Rank (3Yr)",
                    "Excess Performance (5Yr)",
                    "Return Rank (5Yr)",
                    "Sharpe Ratio Rank (5Yr)",
                    "Sortino Ratio Rank (5Yr)",
                    "Expense Ratio Rank",
                    "Investment Style"
                ]

        # Evaluate IPS compliance for each fund
        for block in fund_blocks:
            fund_name = block["Fund Name"]
            fund_type = "Passive" if "bitcoin" in fund_name.lower() else "Active"
            expected_metrics = map_metric_names(fund_type)

            metric_lookup = {m["Metric"]: m["Status"] for m in block["Metrics"]}

            ips_results = []
            for label in expected_metrics[:-1]:  # first 10 metrics
                status = metric_lookup.get(label, "Review")
                ips_results.append("Pass" if status == "Pass" else "Fail")
            ips_results.append("Pass")  # Metric 11 always passes

            fail_count = ips_results.count("Fail")
            if fail_count <= 4:
                overall_status = "Passed IPS Screen"
            elif fail_count == 5:
                overall_status = "Informal Watch (IW)"
            else:
                overall_status = "Formal Watch (FW)"

            # ✅ ADD THIS BLOCK BELOW
            if "step8_results" not in st.session_state:
                st.session_state["step8_results"] = []
            
            st.session_state["step8_results"].append({
                "Fund Name": fund_name,
                "IPS Metrics": ips_results,
                "Overall IPS Status": overall_status
            })

            # Display results
            st.markdown(f"### {fund_name}")
            st.write(f"**Fund Type:** {fund_type}")
            for i, status in enumerate(ips_results, 1):
                st.write(f"- Metric {i}: {status}")
            st.write(f"**Overall IPS Status:** `{overall_status}`")


#------------------------------------------------------------------------------------------------------------------
        
        # === Step 9: Fund Performance: Current vs. Proposed Comparison Section ===
        st.subheader("Step 9: Fund Performance: Current vs. Proposed Comparison Section")

        if fund_perf_pg == "Not found":
            st.error("❌ Could not find the starting page for 'Fund Performance: Current vs. Proposed Comparison'")
        else:
            i = fund_perf_pg - 1  # Convert to 0-based index
            page = pdf.pages[i]
            text = page.extract_text()

            if text and "Fund Performance: Current vs. Proposed Comparison" in text:
                st.markdown(f"**Found section on page {i + 1}**")
                st.text(text[:2000])  # Display first part of the page for verification
            else:
                st.warning(f"⚠️ The text on page {i + 1} does not contain the expected heading. Double-check TOC accuracy.")
                st.text(text[:2000] if text else "No text found on this page.")

#-----------------------------------------------------------------------------------------------------------------

        # === Step 9.4 Redo: Match Investment Option Names & Extract Tickers ===
        st.subheader("Step 9.4: Match Investment Option Names Between Sections")
        
        # Pull fund names from previous Scorecard block
        scorecard_names = [block["Fund Name"] for block in fund_blocks]
        
        # Step 1: Read all pages in Fund Performance section (until heading changes)
        perf_lines = []
        reading = False
        
        with pdfplumber.open(uploaded_file) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue
                if "Fund Performance: Current vs. Proposed Comparison" in text:
                    reading = True
                elif reading and any(kw in text for kw in [
                    "Risk Analysis", "Fund Scorecard", "Definitions & Disclosures", "Table of Contents"
                ]):
                    break
                if reading:
                    perf_lines.extend(text.split("\n"))
        
        # Step 2: Filter candidate lines
        def is_potential_fund_line(line):
            if not (15 <= len(line) <= 90):
                return False
            if any(kw in line.lower() for kw in [
                "matrix", "disclosure", "calendar year", "page", "style", "returns", "mpt statistics"
            ]):
                return False
            return True
        
        clean_lines = [line.strip() for line in perf_lines if is_potential_fund_line(line)]
        
        # Step 3: Fuzzy match each fund name and extract ticker if present
        match_data = []
        for score_name in scorecard_names:
            best_score = 0
            best_line = ""
            best_ticker = ""
            for line in clean_lines:
                score = SequenceMatcher(None, score_name.lower(), line.lower()).ratio()
                if score > best_score:
                    best_score = score
                    ticker_match = re.search(r"\b[A-Z]{5}\b", line)
                    best_ticker = ticker_match.group(0) if ticker_match else ""
                    best_line = re.sub(r"\b[A-Z]{5}\b", "", line).strip()
            match_data.append({
                "Fund Scorecard Name": score_name,
                "Matched Line (Fund Performance)": best_line,
                "Extracted Ticker": best_ticker,
                "Match Score (0–100)": round(best_score * 100),
                "Matched": "✅" if best_score >= 0.60 else "❌"
            })
        
        df_matches = pd.DataFrame(match_data)
        st.dataframe(df_matches)


#--------------------------------------------------------------------------------------------------------------
        
        # === Step 9.5: Extract Fund Ticker and Category ===
        st.subheader("Step 9.5: Extract Fund Ticker and Category")
        
        scorecard_names = [block["Fund Name"] for block in fund_blocks]
        
        # Full category list with proper capitalization
        CATEGORIES = {
            # Equity: Market Cap & Style
            "Large Cap Growth", "Large Cap Value", "Large Cap Blend",
            "Mid Cap Growth", "Mid Cap Value", "Mid Cap Blend",
            "Small Cap Growth", "Small Cap Value", "Small Cap Blend",
            "Micro Cap", "Small Growth",
        
            # Equity: Geography
            "US Equity", "U.S. Equity", "Domestic Equity",
            "International Equity", "Global Equity", "World Stock",
            "Foreign Large Blend", "Foreign Large Growth", "Foreign Large Value",
            "Foreign Small Mid Blend", "Foreign Small Mid Growth", "Foreign Small Mid Value",
            "Emerging Markets", "Diversified Emerging Mkts", "Pacific Asia Ex Japan",
            "Japan Stock", "Europe Stock", "China Region", "Latin America Stock",
        
            # Equity: Sector
            "Health", "Health Care", "Financial", "Financials", "Real Estate", "Global Real Estate",
            "Natural Resources", "Utilities", "Communications", "Technology",
            "Energy", "Consumer Cyclical", "Consumer Defensive",
            "Industrials", "Materials", "Infrastructure",
        
            # Fixed Income
            "Short Term Bond", "Intermediate Term Bond", "Long Term Bond",
            "Ultrashort Bond", "High Yield Bond", "Bank Loan",
            "Corporate Bond", "Government Bond", "Municipal Bond",
            "Inflation Protected Bond", "Inflation-Protected Bond",
            "Global Bond", "Global Bond-USD Hedged", "Emerging Markets Bond",
            "Multi Sector Bond", "Multisector Bond", "Nontraditional Bond",
            "World Bond", "Strategic Income",
            "Treasury Inflation Protected Securities",
            "Intermediate Core Bond", "Intermediate Core-Plus Bond",
        
            # Allocation / Mixed
            "Target Date", "Target Date 2025", "Target Date 2030",
            "Target Date 2035", "Target Date 2040", "Target Date 2045",
            "Target Date 2050", "Target Date 2055", "Target Date 2060",
            "Target-Date 2065+", "Target Retirement Income", "Target-Date Retirement",
            "Target Allocation Aggressive", "Target Allocation Moderate", "Target Allocation Conservative",
            "Balanced Fund", "Flexible Portfolio", "Global Allocation",
            "Tactical Allocation", "Retirement Income Fund",
        
            # Alternatives
            "Alternative Global Macro", "Alternative Long Short Equity", "Alternative Event Driven",
            "Multi Alternative", "Market Neutral", "Bear Market", "Managed Futures",
            "Commodities", "Commodities Broad Basket", "Energy Limited Partnership",
            "Precious Metals", "Real Asset",
        
            # Preservation
            "Money Market", "Stable Value", "Capital Preservation", "Cash Reserves",
            "Ultra Short Term Bond", "Money Market-Taxable"
        }
        category_lookup = {cat.lower(): cat for cat in CATEGORIES}
        known_categories = list(category_lookup.keys())
        
        # === Extract relevant lines and positions from PDF ===
        lines_with_positions = []
        line_to_page_map = []
        reading = False
        
        with pdfplumber.open(uploaded_file) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue
                if "Fund Performance: Current vs. Proposed Comparison" in text:
                    reading = True
                elif reading and any(kw in text for kw in [
                    "Risk Analysis", "Fund Scorecard", "Definitions & Disclosures", "Table of Contents"
                ]):
                    break
                if reading:
                    words = page.extract_words()
                    line_map = {}
                    for w in words:
                        y0 = round(w['top'], 1)
                        if y0 not in line_map:
                            line_map[y0] = []
                        line_map[y0].append(w['text'])
                    for y in sorted(line_map.keys()):
                        line_text = " ".join(line_map[y])
                        lines_with_positions.append((y, line_text))
                        line_to_page_map.append(i)
        
        # === Match fund names, extract ticker and category ===
        results = []
        for name in scorecard_names:
            best_score = 0
            best_idx = -1
            best_line = ""
            best_ticker = ""
        
            for i, (y, line) in enumerate(lines_with_positions):
                score = SequenceMatcher(None, name.lower(), line.lower()).ratio()
                if score > best_score:
                    best_score = score
                    best_idx = i
                    ticker_match = re.search(r"\b[A-Z]{5}\b", line)
                    best_ticker = ticker_match.group(0) if ticker_match else ""
                    best_line = re.sub(r"\b[A-Z]{5}\b", "", line).strip()
        
            # Scan upward for category match
            category = ""
            for j in range(best_idx - 1, max(0, best_idx - 20), -1):
                cand = lines_with_positions[j][1].strip().lower()
                match = get_close_matches(cand, known_categories, n=1, cutoff=0.8)
                if match:
                    category = category_lookup[match[0]]
                    break
        
            results.append({
                "Fund Scorecard Name": name,
                "Matched Name": best_line,
                "Ticker": best_ticker,
                "Category": category
            })
        
        df_95 = pd.DataFrame(results)
        st.dataframe(df_95)

#--------------------------------------------------------------------------------------------------------------

        # === Step 10: Output Combined Table (IPS Summary) ===
        st.subheader("Step 10: IPS Summary Table")
        
        # Rebuild fund_blocks with ticker + IPS metrics + status
        final_fund_blocks = {}
        for i, block in enumerate(fund_blocks):
            name = block["Fund Name"]
            fund_metrics = [m["Status"] for m in block["Metrics"]]
            fund_status = "Unknown"
        
            # Match IPS Status
            for line in st.session_state.get("step8_results", []):
                if line.get("Fund Name") == name:
                    fund_status = line.get("Overall IPS Status", "Unknown")
                    break
        
            # Match Ticker from df_95
            ticker_row = df_95[df_95["Fund Scorecard Name"] == name]
            ticker_val = ticker_row["Ticker"].values[0] if not ticker_row.empty else "N/A"
        
            final_fund_blocks[name] = {
                "Ticker": ticker_val,
                "Metrics": fund_metrics[:11],  # Only IPS Metrics (first 11)
                "IPS Status": fund_status
            }
        
        # Build Table
        def generate_step10_table(fund_blocks, quarter_label):
            data = []
            for fund_name, info in fund_blocks.items():
                row = {
                    "Investment Option": fund_name,
                    "Ticker": info.get("Ticker", "N/A"),
                    "Time Period": quarter_label,
                    "Plan Assets": "$"
                }
                for i, metric in enumerate(info.get("Metrics", []), start=1):
                    row[str(i)] = metric
                row["IPS Status"] = info.get("IPS Status", "N/A")
                data.append(row)
        
            columns = (
                ["Investment Option", "Ticker", "Time Period", "Plan Assets"] +
                [str(i) for i in range(1, 12)] + ["IPS Status"]
            )
            return pd.DataFrame(data, columns=columns)
        
        # Show Table
        summary_df = generate_step10_table(final_fund_blocks, quarter)
        def color_cells(val):
            if val == "Pass":
                return "background-color: #c6e7b5; color: black;"
            elif val == "Fail":
                return "background-color: #e09385; color: black;"
            return ""
        
        def color_status(val):
            if "Passed" in val:
                return "background-color: #8ea781; color: black;"
            elif "Informal" in val:
                return "background-color: #e0b485; color: black;"
            elif "Formal" in val:
                return "background-color: #e09385; color: black;"
            return ""
        
        styled_df = summary_df.style.applymap(color_cells, subset=[str(i) for i in range(1, 12)])
        styled_df = styled_df.applymap(color_status, subset=["IPS Status"])
        
        st.dataframe(styled_df, use_container_width=True)
        
        # Download Button
        csv = summary_df.to_csv(index=False).encode("utf-8")
        st.download_button("Download IPS Summary CSV", data=csv, file_name="ips_summary.csv", mime="text/csv")

#--------------------------------------------------------------------------------------------------------------




         # === Step 10: Output Combined Table (IPS Summary) ===
        st.subheader("Step 11: Excel Table")

        def export_step11_excel(summary_df, file_path="Step11_Final_With10.xlsx"):
            # === Sheet 1 Content ===
            ips_criteria = [
                "Manager Tenure ≥ 3 years",
                "*3-Year Performance > Benchmark / +3-Year R² > 95%",
                "3-Year Performance > 50% of Peers",
                "3-Year Sharpe Ratio > 50% of Peers",
                "*3-Year Sortino Ratio > 50% of Peers / +3-Year Tracking Error < 90% of Peers",
                "*5-Year Performance > Benchmark / +5-Year R² > 95%",
                "5-Year Performance > 50% of Peers",
                "5-Year Sharpe Ratio > 50% of Peers",
                "*5-Year Sortino Ratio > 50% of Peers / +5-Year Tracking Error < 90% of Peers",
                "Expense Ratio < 50% of Peers",
                "Investment Style aligns with fund objectives"
            ]
        
            criteria_thresholds = [
                "Manager Tenure - Portfolio manager or management team must have managed this product for at least 3 years.",
                "Excess Performance (3Yr) - The fund must outperform its benchmark over the trailing 3-year period.",
                "Excess Performance (5Yr) - The fund must outperform its benchmark over the trailing 5-year period.",
                "Peer Return Rank (3Yr) - The fund's Return Rank must be in the top 50% of its peer group over the trailing 3-year period.",
                "Peer Return Rank (5Yr) - The fund's Return Rank must be in the top 50% of its peer group over the trailing 5-year period.",
                "Expense Ratio Rank - The fund's Expense Ratio must be in the top 50% of its peer group.",
                "Sharpe Ratio Rank (3Yr) - The fund's Sharpe Ratio Rank must be in the top 50% of its peer group over the trailing 3-year period.",
                "Sharpe Ratio Rank (5Yr) - The fund's Sharpe Ratio Rank must be in the top 50% of its peer group over the trailing 5-year period.",
                "R-Squared (3Yr) - The fund's Benchmark R-Squared must be greater than 95% over the trailing 3-year period.",
                "R-Squared (5Yr) - The fund's Benchmark R-Squared must be greater than 95% over the trailing 5-year period.",
                "Sortino Ratio Rank (3Yr) - The fund's Sortino Ratio Rank must be in the top 50% of its peer group over the trailing 3-year period.",
                "Sortino Ratio Rank (5Yr) - The fund's Sortino Ratio Rank must be in the top 50% of its peer group over the trailing 5-year period.",
                "Tracking Error Rank (3Yr) - The fund's Tracking Error Rank must be in the top 10% of its peer group over the trailing 3-year period.",
                "Tracking Error Rank (5Yr) - The fund's Tracking Error Rank must be in the top 10% of its peer group over the trailing 5-year period.",
            ]
        
            # === Create Workbook ===
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "IPS & Scorecard Criteria"
            ws1.append(["IPS Investment Criteria"])
            for line in ips_criteria:
                ws1.append([line])
            ws1.append([])
            ws1.append(["Criteria Threshold (14 Metrics)"])
            for line in criteria_thresholds:
                ws1.append([line])
        
            # === Sheet 2: IPS Summary Table ===
            ws2 = wb.create_sheet("IPS Summary Table")
            for row in dataframe_to_rows(summary_df, index=False, header=True):
                ws2.append(row)
        
            # === Styling ===
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_green = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid")
            status_red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
            status_orange = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
            white_font = Font(color="FFFFFF")
        
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                for cell in row[4:-1]:
                    if cell.value == "Pass":
                        cell.fill = green_fill
                    elif cell.value == "Fail":
                        cell.fill = red_fill
                status_cell = row[-1]
                if "Passed" in status_cell.value:
                    status_cell.fill = status_green
                    status_cell.font = white_font
                elif "Informal" in status_cell.value:
                    status_cell.fill = status_orange
                    status_cell.font = white_font
                elif "Formal" in status_cell.value:
                    status_cell.fill = status_red
                    status_cell.font = white_font
        
            # === Save File ===
            wb.save(file_path)
            return file_path

        if st.button("Download Excel Output"):
            excel_path = export_step11_excel(summary_df)
            with open(excel_path, "rb") as f:
                st.download_button("Download Step 11 Excel", data=f, file_name="Step11_IPS_Export.xlsx")
