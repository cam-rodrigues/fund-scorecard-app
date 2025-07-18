import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# --- IPS Metric Names ---
IPS_METRICS = [
    "Manager Tenure",
    "3Y: Benchmark/R²",
    "3Y: Peer Rank",
    "3Y: Sharpe",
    "3Y: Sortino/TE",
    "5Y: Benchmark/R²",
    "5Y: Peer Rank",
    "5Y: Sharpe",
    "5Y: Sortino/TE",
    "Expense Ratio",
    "Investment Style"
]

# --- Color Styles ---
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
WHITE_TEXT = Font(color="FFFFFF")

# --- Utility: Extract text safely ---
def safe_extract(page):
    try:
        return page.extract_text()
    except:
        return ""

# --- IPS Pass/Fail logic ---
def evaluate_ips_status(metrics, fund_type):
    passes = []
    for idx, metric in enumerate(metrics[:10]):
        if metric == "Pass":
            passes.append(True)
        else:
            passes.append(False)
    passes.append(True)  # Investment Style is always Pass

    fails = 11 - sum(passes)
    if fails <= 4:
        return "Passed IPS Screen", GREEN
    elif fails == 5:
        return "Informal Watch", ORANGE
    else:
        return "Formal Watch", RED

# --- Main App ---
def run():
    st.set_page_config(page_title="IPS Investment Criteria", layout="wide")
    st.title("IPS Investment Criteria Screening")

    uploaded_file = st.file_uploader("Upload an MPI-style PDF", type=["pdf"])
    if not uploaded_file:
        st.stop()

    # Extract PDF content
    with pdfplumber.open(uploaded_file) as pdf:
        raw_text = [safe_extract(p) for p in pdf.pages]
        page1 = raw_text[0] if raw_text else ""
        toc_page = raw_text[1] if len(raw_text) > 1 else ""

        # --- Step 2: Extract Time Period ---
        time_period = None
        for line in page1.split("\n"):
            if re.search(r"\b(3/31|6/20|9/30|12/31)/20\d{2}\b", line):
                time_period = line.strip()
                break

        # --- Step 3: Page 1 Info ---
        total_options = None
        prepared_for = None
        for line in page1.split("\n"):
            if "Total Options" in line:
                total_options = re.search(r"Total Options: ?(\d+)", line)
                total_options = int(total_options.group(1)) if total_options else None
            if "Prepared For" in line:
                idx = page1.split("\n").index(line)
                prepared_for = page1.split("\n")[idx + 1].strip()

        # --- Step 4: Locate section pages from ToC ---
        def find_section_page(section_name):
            for line in toc_page.split("\n"):
                if section_name in line:
                    match = re.search(r"\s(\d+)$", line.strip())
                    return int(match.group(1)) - 1 if match else None
            return None

        perf_pg = find_section_page("Fund Performance: Current vs. Proposed Comparison")
        scorecard_pg = find_section_page("Fund Scorecard")

        if perf_pg is None or scorecard_pg is None:
            st.error("Failed to locate section pages.")
            st.stop()

        # --- Step 5: Extract Fund Scorecard Section ---
        fund_data = []
        fund_name = None
        current_metrics = []
        fund_lines = "\n".join(raw_text[scorecard_pg:])

        for line in fund_lines.split("\n"):
            if re.match(r"^[A-Z].+\b(Fund|ETF)\b", line):
                if fund_name and current_metrics:
                    fund_data.append((fund_name, current_metrics))
                fund_name = line.strip()
                current_metrics = []
            elif any(metric in line for metric in IPS_METRICS):
                if "Pass" in line:
                    current_metrics.append("Pass")
                elif "Review" in line:
                    current_metrics.append("Fail")

        if fund_name and current_metrics:
            fund_data.append((fund_name, current_metrics))

        # --- Step 6: Match to Performance Table ---
        ticker_lookup = {}
        category_lookup = {}
        perf_lines = "\n".join(raw_text[perf_pg:]).split("\n")
        for i, line in enumerate(perf_lines):
            if re.match(r"^[A-Z]{5}$", line.strip()):
                fund = perf_lines[i - 1].strip()
                category = perf_lines[i - 2].strip()
                ticker = line.strip()
                ticker_lookup[fund] = ticker
                category_lookup[fund] = category

        # --- Step 7: Build Output Table ---
        wb = Workbook()
        ws = wb.active
        headers = [
            "Investment Option", "Category", "Ticker", "Time Period", "Plan Assets"
        ] + [f"{i+1}" for i in range(11)] + ["IPS Status"]
        ws.append(headers)

        for fund, metrics in fund_data:
            ticker = ticker_lookup.get(fund, "N/A")
            category = category_lookup.get(fund, "N/A")
            fund_type = "Passive" if "Bitcoin" in fund else "Active"
            status, status_fill = evaluate_ips_status(metrics, fund_type)

            row = [fund, category, ticker, time_period, "$"] + metrics + [status]
            ws.append(row)
            for i, val in enumerate(metrics):
                cell = ws.cell(row=ws.max_row, column=6 + i)
                cell.fill = GREEN if val == "Pass" else RED

            # Format IPS Status Cell
            status_cell = ws.cell(row=ws.max_row, column=len(headers))
            status_cell.fill = status_fill
            status_cell.font = WHITE_TEXT

        # --- Step 8: Display & Download ---
        st.success("IPS Evaluation Complete")
        st.dataframe(pd.DataFrame([r[:5] + r[5:-1] + [r[-1]] for r in ws.iter_rows(min_row=2, values_only=True)],
                                  columns=headers))

        buffer = BytesIO()
        wb.save(buffer)
        st.download_button("Download Excel", buffer.getvalue(), file_name="IPS_Evaluation.xlsx")

if __name__ == "__main__":
    run()
