from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

def update_excel_with_template(excel_bytes, sheet_name, match_df):
    try:
        wb = load_workbook(excel_bytes)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        ws = wb[sheet_name]

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

        for i, row in enumerate(match_df.itertuples(index=False), start=2):
            ws.cell(row=i, column=1, value=row._1)  # Fund Name (Raw)
            ws.cell(row=i, column=2, value=row._2)  # Matched Option
            ws.cell(row=i, column=3, value=row._3)  # Score
            status_cell = ws.cell(row=i, column=4, value=row._4)  # Status
            status_cell.fill = green_fill if row._4 == "Pass" else red_fill

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        raise RuntimeError(f"Excel update failed: {e}")
