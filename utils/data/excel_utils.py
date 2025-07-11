from openpyxl.styles import PatternFill
import openpyxl

def update_excel_with_template(excel_file, sheet_name, match_data, fund_col, status_col):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb[sheet_name]

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        match_dict = {}
        for item in match_data:
            if isinstance(item, (list, tuple)) and len(item) == 2:
                fund_name, status = item
                match_dict[fund_name.strip().lower()] = status
            else:
                print("⚠️ Skipping invalid row in match_data:", item)

        row = fund_col[1] + 1  # Start below header
        while ws.cell(row=row, column=fund_col[0]).value:
            fund = str(ws.cell(row=row, column=fund_col[0]).value).strip().lower()
            if fund in match_dict:
                status = match_dict[fund]
                status_cell = ws.cell(row=row, column=status_col[0])
                status_cell.value = status
                status_cell.fill = PatternFill()  # clear fill

                if status == "Pass":
                    status_cell.fill = green_fill
                elif status == "Review":
                    status_cell.fill = red_fill
            row += 1

        wb.save(excel_file)
        return True, "✅ Excel updated successfully."

    except Exception as e:
        return False, f"❌ Failed to update Excel: {e}"
