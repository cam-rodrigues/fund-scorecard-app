import streamlit as st
import io
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rapidfuzz import process, fuzz
import base64
import string
from datetime import datetime
import gc

st.set_page_config(page_title="Fund Scorecard • FidSync", layout="wide")

st.markdown("""
    <style>
        .block-container {
            padding-top: 2rem;
        }
        .stDownloadButton>button {
            color: white;
            background-color: #003865;
            border-radius: 6px;
            font-weight: 500;
        }
        .stButton>button {
            border-radius: 6px;
        }
        .status-pass {
            background-color: #C6EFCE;
            color: #006100;
            padding: 2px 8px;
            border-radius: 5px;
            font-weight: 500;
        }
        .status-fail {
            background-color: #FFC7CE;
            color: #9C0006;
            padding: 2px 8px;
            border-radius: 5px;
            font-weight: 500;
        }
    </style>
""", unsafe_allow_html=True)

GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

def normalize_name(name):
    name = name.lower().translate(str.maketrans('', '', string.punctuation))
    return " ".join(name.split())

def extract_fund_status(pdf_bytes, start_page, end_page):
    fund_status = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num in range(start_page - 1, end_page):
            if page_num >= len(pdf.pages):
                continue
            text = pdf.pages[page_num].extract_text() or ""
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if "manager tenure" in line.lower() and i > 0:
                    fund_line = lines[i - 1].strip()
                    if "Fund Meets Watchlist Criteria." in fund_line:
                        name = fund_line.split("Fund Meets Watchlist Criteria.")[0].strip()
                        fund_status[normalize_name(name)] = "Pass"
                    elif "Fund has been placed on watchlist" in fund_line:
                        name = fund_line.split("Fund has been placed on watchlist")[0].strip()
                        fund_status[normalize_name(name)] = "Fail"
    return fund_status

def update_excel_with_status(pdf_bytes, excel_bytes, sheet_name, status_col, start_row, fund_names, start_page, end_page, dry_run=False):
    fund_status_map = extract_fund_status(pdf_bytes, start_page, end_page)
    pdf_names = list(fund_status_map.keys())
    wb = load_workbook(io.BytesIO(excel_bytes))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]

    updated_count = 0
    match_log = []

    for i, fund_name in enumerate(fund_names):
        normalized = normalize_name(fund_name)
        match = process.extractOne(normalized, pdf_names, scorer=fuzz.token_sort_ratio)
        row = start_row + i
        if match and match[1] >= 70:
            matched_name, score = match
            status = fund_status_map.get(matched_name)
            if not dry_run:
                cell = ws[f"{status_col}{row}"]
                if cell.data_type != 'f':
                    cell.value = status
                    cell.fill = GREEN_FILL if status == "Pass" else RED_FILL if status == "Fail" else PatternFill()
                    cell.number_format = 'General'
                    updated_count += 1
            match_log.append((fund_name, matched_name, score, status))
        else:
            match_log.append((fund_name, "No match", 0, "N/A"))

    out_bytes = io.BytesIO()
    wb.save(out_bytes)
    out_bytes.seek(0)
    return out_bytes, updated_count, match_log

st.title("Fund Scorecard Status Tool")

with st.form("upload_form"):
    st.header("📄 Upload & Settings")
    pdf_file = st.file_uploader("Upload Fund Scorecard PDF", type=["pdf"])
    excel_file = st.file_uploader("Upload Excel Workbook", type=["xlsx", "xlsm"])

    col1, col2, col3 = st.columns(3)
    sheet_name = col1.text_input("Excel Sheet Name")
    status_col = col2.text_input("Starting Column Letter").strip().upper()
    start_row = col3.number_input("Starting Row Number", min_value=1)

    col4, col5 = st.columns(2)
    start_page = col4.number_input("Start Page in PDF", min_value=1)
    end_page = col5.number_input("End Page in PDF", min_value=1)

    fund_names_input = st.text_area("Investment Option Names (One Per Line)", height=200)
    dry_run = st.checkbox("Dry Run (preview only, don't update Excel)", value=False)

    submitted = st.form_submit_button("Run Status Update")

if submitted:
    if not pdf_file or not excel_file:
        st.warning("Please upload both PDF and Excel files.")
    elif start_page > end_page:
        st.error("Start Page must be less than or equal to End Page.")
    elif not fund_names_input.strip():
        st.warning("Please enter investment option names.")
    else:
        with st.spinner("Processing..."):
            try:
                fund_names = [line.strip() for line in fund_names_input.strip().splitlines() if line.strip()]
                pdf_bytes = pdf_file.read()
                excel_bytes = excel_file.read()

                updated_excel, count, match_log = update_excel_with_status(
                    pdf_bytes, excel_bytes, sheet_name, status_col,
                    start_row, fund_names, start_page, end_page,
                    dry_run
                )

                st.success(f"Updated {count} row(s)" if not dry_run else "Dry run complete. No changes made.")

                if not dry_run:
                    timestamp = datetime.now().strftime("%b-%d-%Y_%H%M")
                    b64 = base64.b64encode(updated_excel.getvalue()).decode()
                    link = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="Updated_Status_{timestamp}.xlsx">📥 Download Updated Excel</a>'
                    st.markdown(link, unsafe_allow_html=True)

                with st.expander("Match Log", expanded=True):
                    st.markdown("**Matched Results:**")
                    styled_log = []
                    for name, matched, score, status in match_log:
                        badge = ""
                        if status == "Pass":
                            badge = f"<span class='status-pass'>{status}</span>"
                        elif status == "Fail":
                            badge = f"<span class='status-fail'>{status}</span>"
                        styled_log.append([name, matched, score, badge])
                    df = pd.DataFrame(styled_log, columns=["Input Name", "Matched Name", "Match Score", "Status"])
                    st.write(df.to_html(escape=False, index=False), unsafe_allow_html=True)

                    csv_buffer = io.StringIO()
                    pd.DataFrame(match_log, columns=["Input Name", "Matched Name", "Match Score", "Status"]).to_csv(csv_buffer, index=False)
                    csv_b64 = base64.b64encode(csv_buffer.getvalue().encode()).decode()
                    csv_link = f'<a href="data:file/csv;base64,{csv_b64}" download="match_log.csv">📤 Download Match Log CSV</a>'
                    st.markdown(csv_link, unsafe_allow_html=True)

                del pdf_bytes, excel_bytes, updated_excel
                gc.collect()

            except Exception as e:
                st.error("Something went wrong.")
                st.exception(e)
